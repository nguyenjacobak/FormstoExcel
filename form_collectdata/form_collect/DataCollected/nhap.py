import pandas as pd
import numpy as np
from openpyxl import load_workbook
def process_final_new():
    # Đọc file Excel
    final_new_path = r"FormstoExcel\form_collectdata\form_collect\DataCollected\final_new.xlsx"
    tong_hop_diem_path = r"FormstoExcel\form_collectdata\form_collect\DataCollected\TongHopDiem1.xlsx"
    
    df_final_new = pd.read_excel(final_new_path)
    df_tong_hop_diem = pd.read_excel(tong_hop_diem_path)
    
    # Điền giá trị NaN bằng 0
    df_final_new = df_final_new.fillna(0)
    
    # Danh sách các cột cần lấy dữ liệu
    columns_to_extract = [
        'Họ và tên', 'Mã sinh viên', 'Lớp',
        'HDCM_uv1_C3', 'HDCM_uv1_C4', 'HDCM_uv1_C5', 'HDCM_uv1_C6',
        'HDCM_uv2_C3', 'HDCM_uv2_C4', 'HDCM_uv2_C5', 'HDCM_uv2_C6',
        'HDCM_uv3_C3', 'HDCM_uv3_C4', 'HDCM_uv3_C5', 'HDCM_uv3_C6',
        'HDCM_uv4_C3', 'HDCM_uv4_C4', 'HDCM_uv4_C5', 'HDCM_uv4_C6',
        'HDCM_uv5_C3', 'HDCM_uv5_C4', 'HDCM_uv5_C5', 'HDCM_uv5_C6',
        'CBHD_1_C1', 'CBHD_1_C5',
        'CBHD_2_C2', 'CBHD_2_C3', 'CBHD_2_C5',
        'CBHD_3_C2', 'CBHD_3_C3', 'CBHD_3_C4', 'CBHD_3_C6',
        'CBPB_C2', 'CBPB_C3', 'CBPB_C4', 'CBPB_C6',
        'o1TB', 'o2TB', 'o3TB', 'o4TB', 'o5TB', 'o6TB', 'gpa'
    ]
    
    # Hàm tính trung bình của các cột con
    def calculate_average(df, cols):
        existing_cols = [col for col in cols if col in df.columns]
        if existing_cols:
            return df[existing_cols].mean(axis=1)
        else:
            return pd.Series([0] * len(df))
    
    # Tạo DataFrame mới với các cột cần thiết
    new_data = pd.DataFrame()
    new_data['Họ và tên'] = df_final_new['Họ và tên']
    new_data['Mã sinh viên'] = df_final_new['Mã sinh viên']
    new_data['Lớp'] = df_final_new['Lớp']
    
    # Tính trung bình cho các cột HDCM_uv
    for i in range(1, 6):
        for j in range(3, 7):
            col_name = f'HDCM_uv{i}_C{j}'
            sub_cols = [f'HDCM_uv{i}_C{j}.{k}' for k in range(1, 5) if f'HDCM_uv{i}_C{j}.{k}' in df_final_new.columns]
            new_data[col_name] = calculate_average(df_final_new, sub_cols)
    
    # Tính trung bình cho các cột CBHD
    for i in range(1, 4):
        for j in [1, 2, 3, 4, 5, 6]:
            col_name = f'CBHD_{i}_C{j}'
            sub_cols = [f'CBHD_{i}_C{j}.{k}' for k in range(1, 5) if f'CBHD_{i}_C{j}.{k}' in df_final_new.columns] 
            new_data[col_name] = calculate_average(df_final_new, sub_cols)
    
    # Tính trung bình cho các cột CBPB
    for j in [2, 3, 4, 6]:
        col_name = f'CBPB_C{j}'
        sub_cols = [f'CBPB_C{j}.{k}' for k in range(1, 5) if f'CBPB_C{j}.{k}' in df_final_new.columns]
        new_data[col_name] = calculate_average(df_final_new, sub_cols)
    
    # Tính trung bình cho các cột o1TB, o2TB, ..., o6TB
    for i in range(1, 7):
        col_name = f'o{i}TB'
        sub_cols = [f'HDCM_uv{j}_C{i}' for j in range(1, 6) if f'HDCM_uv{j}_C{i}' in columns_to_extract] + \
           [f'CBHD_{j}_C{i}' for j in range(1, 4) if f'CBHD_{j}_C{i}' in columns_to_extract] + \
           [f'CBPB_C{i}' for j in range(1, 2) if f'CBPB_C{i}' in columns_to_extract]
        new_data[col_name] = calculate_average(new_data, sub_cols)
        # print(sub_cols)
        # print(df_final_new.columns)

    # Tính trung bình cho cột gpa
    new_data['gpa'] = new_data[[f'o{i}TB' for i in range(1, 7)]].mean(axis=1)
    
    # Chỉ giữ lại các cột cần thiết
    new_data = new_data[columns_to_extract]

    # Xóa dữ liệu cũ từ hàng số 4 trở đi
    workbook = load_workbook(tong_hop_diem_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

    # Lưu workbook sau khi xóa dữ liệu cũ
    workbook.save(tong_hop_diem_path)
    
    # Thêm dữ liệu mới vào file TongHopDiem1.xlsx
    with pd.ExcelWriter(tong_hop_diem_path, mode='a', if_sheet_exists='overlay') as writer:
        new_data.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
    
    print("Dữ liệu đã được thêm vào file TongHopDiem1.xlsx")

# Gọi hàm để xử lý dữ liệu
process_final_new()