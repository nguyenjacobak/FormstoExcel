import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook

def process_final_new_baocao2():
    # Đọc file Excel
    final_new_path = r"FormstoExcel\form_collectdata\form_collect\DataCollected\final_new.xlsx"
    
    df_final_new = pd.read_excel(final_new_path)
    
    # Điền giá trị NaN bằng 0
    df_final_new = df_final_new.fillna(0)
    
    # Danh sách các cột cần lấy dữ liệu
    columns_to_extract = [
        'TT',
        'Họ và tên', 'Mã sinh viên', 'Lớp',
        'CBHD_1_C1', 'CBHD_1_C5', 'CBHD_1_gpa',
        'CBHD_2_C2', 'CBHD_2_C3', 'CBHD_2_C5', 'CBHD_2_gpa',
        'o1TB', 'o2TB', 'o3TB', 'o4TB', 'o5TB', 'o6TB', 'oGPA',
        'o2CK', 'o3CK', 'o4CK', 'o5CK', 'o6CK', 'GPA_CK',
        'GPA_tong'
    ]
    
    # Hàm tính trung bình của các cột con
    def calculate_average(df, cols):
        existing_cols = [col for col in cols if col in df.columns]
        if existing_cols:
            return df[existing_cols].mean(axis=1)
        else:
            return pd.Series([0] * len(df))
    
    # Tạo DataFrame mới với các cột cần thiết
    new_data = pd.DataFrame()
    new_data['TT'] = range(1, len(df_final_new) + 1)
    new_data['Họ và tên'] = df_final_new['Họ và tên']
    new_data['Mã sinh viên'] = df_final_new['Mã sinh viên']
    new_data['Lớp'] = df_final_new['Lớp']
    
    # Tính trung bình cho các cột CBHD_1_C1, CBHD_1_C5, CBHD_2_C2, CBHD_2_C3, CBHD_2_C5
    for i in range(1, 3):
        for j in [1, 2, 3, 5]:
            if i == 1 and j == 2:
                continue
            col_name = f'CBHD_{i}_C{j}'
            sub_cols = [f'CBHD_{i}_C{j}.{k}' for k in range(1, 5) if f'CBHD_{i}_C{j}.{k}' in df_final_new.columns]
            new_data[col_name] = calculate_average(df_final_new, sub_cols)
    
    # Thêm các cột CBHD_1_gpa và CBHD_2_gpa
    new_data['CBHD_1_gpa'] = df_final_new['CBHD_1_gpa']
    new_data['CBHD_2_gpa'] = df_final_new['CBHD_2_gpa']
    
    # Tính trung bình cho các cột o1TB, o2TB, ..., o6TB
    for i in range(1, 7):
        col_name = f'o{i}TB'
        sub_cols = [f'HDCM_uv{j}_C{i}.{k}' for j in range(1, 6) for k in range(1, 5) if f'HDCM_uv{j}_C{i}.{k}' in df_final_new.columns] + \
                   [f'CBHD_{j}_C{i}.{k}' for j in range(1, 4) for k in range(1, 5) if f'CBHD_{j}_C{i}.{k}' in df_final_new.columns] + \
                   [f'CBPB_C{i}.{k}' for k in range(1, 5) if f'CBPB_C{i}.{k}' in df_final_new.columns]
        new_data[col_name] = calculate_average(df_final_new, sub_cols)
    
    # Tính trung bình cho cột oGPA
    new_data['oGPA'] = new_data[[f'o{i}TB' for i in range(1, 7)]].mean(axis=1)
    
    # Tính trung bình cho các cột o2CK, o3CK, ..., o6CK
    for i in range(2, 7):
        col_name = f'o{i}CK'
        sub_cols = [f'HDCM_uv{j}_C{i}.{k}' for j in range(1, 6) for k in range(1, 5) if f'HDCM_uv{j}_C{i}.{k}' in df_final_new.columns] + \
                   [f'CBHD_{j}_C{i}.{k}' for j in range(3,4) for k in range(1, 5) if f'CBHD_{j}_C{i}.{k}' in df_final_new.columns] + \
                   [f'CBPB_C{i}.{k}' for k in range(1, 5) if f'CBPB_C{i}.{k}' in df_final_new.columns]
        new_data[col_name] = calculate_average(df_final_new, sub_cols)
    
    # Tính trung bình cho cột GPA_CK
    new_data['GPA_CK'] = new_data[[f'o{i}CK' for i in range(2, 7)]].mean(axis=1)
    
    # Tính trung bình cho cột GPA_tong
    new_data['GPA_tong'] = new_data['CBHD_1_gpa'] * 0.1 + new_data['CBHD_2_gpa'] * 0.2 + new_data['oGPA'] * 0.35 + new_data['GPA_CK'] * 0.35
    
    # Chỉ giữ lại các cột cần thiết
    new_data = new_data[columns_to_extract]

    # Đường dẫn đến file TongHopDiem2.xlsx
    tong_hop_diem2_path = r"FormstoExcel\form_collectdata\form_collect\DataCollected\TongHopDiem2.xlsx"

    # Xóa dữ liệu cũ từ hàng số 4 trở đi và thêm dữ liệu mới
    workbook = load_workbook(tong_hop_diem2_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

    # Lưu workbook sau khi xóa dữ liệu cũ
    workbook.save(tong_hop_diem2_path)

    # Thêm dữ liệu mới vào file TongHopDiem2.xlsx
    with pd.ExcelWriter(tong_hop_diem2_path, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
        new_data.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=3)

    print("Dữ liệu đã được thêm vào file TongHopDiem2.xlsx")

# Gọi hàm để xử lý dữ liệu và lấy kết quả
process_final_new()

