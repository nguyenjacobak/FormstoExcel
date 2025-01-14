from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook,load_workbook
import os
import pandas as pd
from django.http import JsonResponse
import json
from django.shortcuts import redirect
import re
import numpy as np

# Create your views here.
def get_students_view(request):
    data = request.body.decode('utf-8')
    project_name = json.loads(data).get('project_name', '')
    students = get_students_by_project_name(project_name)
    return JsonResponse(students, safe=False)

def get_all_students_view(request):
    students = get_all_students()
    return JsonResponse(students, safe=False)

def get_all_councils_view(request):
    councils = get_all_councils()
    return JsonResponse(councils, safe=False)

def get_students_by_project_name(project_name):
    index = 0
    for i in range(len(project_name)):
        if project_name[i] == '(':
            index = i
    
    list_msv = project_name[index+1 : -1].split(' - ')
    project_name = project_name[:index].strip()
    file_path = os.path.join('DataBase', 'db.xlsx')
    df = pd.read_excel(file_path, sheet_name="Danh sách các đồ án ", skiprows=12)
    df = df.fillna(method='ffill')

    students = []
    col = 'Tên đề tài đồ án/ khóa luận tốt nghiệp'
    df_students = df[df[col].str.contains(project_name, case=False, regex=False ) & df['Mã sinh viên'].isin(list_msv)]
    first_students = list(df_students['Họ và tên'])
    second_students = list(df_students['Unnamed: 3'])
    fullname_students = [f"{first} {second}" for first, second in zip(first_students, second_students)]
    msv_students = list(df_students['Mã sinh viên'])
    day_of_birth_students = list(df_students['Năm sinh'])
    class_students = list(df_students['Lớp'])
    for i in range(len(fullname_students)):
        students.append({
            'fullname': fullname_students[i],
            'msv': msv_students[i],
            'day_of_birth': day_of_birth_students[i],
            'class': class_students[i]
        })
    return students

def get_lecturers():
    file_path = os.path.join('DataBase', 'db.xlsx')
    df = pd.read_excel(file_path, sheet_name="DS Hội đồng_DN", skiprows=2)
    df = df.drop(columns = ['Unnamed: 4'])
    df = df.dropna(ignore_index=True)
    df = df[df['Họ và tên'] != 'Họ và tên']
    lecture_in_sheet1 = list(df['Họ và tên'].values)

    lecturers = set()
    # Get hoi dong
    for lecture in lecture_in_sheet1:
        lecture =  lecture.replace("TS", "")
        lecturers.add(lecture.split(". ")[1].strip())

    # Get phan bien
    df2 = pd.read_excel(file_path, sheet_name="DS SV các Hội đồng phân PB", skiprows=1)
    df2 = df2.drop(columns = ['Unnamed: 0'])
    df2.loc[1, 'Unnamed: 11'] = df2.loc[0, 'Unnamed: 11']
    df2.loc[1, 'Unnamed: 12'] = df2.loc[0, 'Unnamed: 12']
    df2.loc[1, 'Unnamed: 3'] = 'Tên'
    df2.columns = df2.iloc[1]
    df2 = df2.dropna(how='all', ignore_index=True)
    for i in range(len(df2)):
        msv = df2.loc[i, 'Mã sinh viên']
        if 'Hội đồng' in msv or msv == 'Mã sinh viên':
            continue
        opponent = df2.loc[i, 'Phản biện  (Khoa)'].strip() if not pd.isna(df2.loc[i, 'Phản biện  (Khoa)']) else ''
        if opponent != '':
            opponent = opponent.replace("TS", "")
            lecturers.add(opponent.split(". ")[1].strip())

    # Get thay co huong dan
    df3 = pd.read_excel(file_path, sheet_name="Danh sách các thầy cô")
    lecturers.add(list(df3.columns)[0])
    list_lecturers_in_sheet4 = df3.values.tolist()
    for i in list_lecturers_in_sheet4:
        pattern = r'()'
        if not bool(re.search(pattern, i[0])):
            lecturers.add(i[0])

    lecturesList = list(lecturers)
    return sorted(lecturesList, key=lambda lecture: lecture.split()[-1])


def get_projects_by_lecture_and_type(lecturer_name, project_type):
    file_path = os.path.join('DataBase', 'db.xlsx')
    df = pd.read_excel(file_path, sheet_name="Danh sách các đồ án ", skiprows=12)
    df = df.fillna(method='ffill')

    projects = []
    cols = ['Tên đề tài đồ án/ khóa luận tốt nghiệp', 'Giáo viên hướng dẫn', 'Làm đồ án/Học phần TTTN']
    for col in cols:
        if col not in df.columns:
            return projects

    projects = df[df['Giáo viên hướng dẫn'].str.contains(lecturer_name, case=False) & df['Làm đồ án/Học phần TTTN'].str.contains(project_type, case=False)]['Tên đề tài đồ án/ khóa luận tốt nghiệp']

    projects = list(set(projects))
    return projects

def find_student_by_council_and_group_id(council_id=None, group_id=None):
    students = []
    if group_id is None or council_id is None:
        return students
    file_path = os.path.join('DataBase', 'db.xlsx')
    df2 = pd.read_excel(file_path, sheet_name="DS SV các Hội đồng phân PB", skiprows=1)
    df2 = df2.drop(columns = ['Unnamed: 0'])
    df2.loc[1, 'Unnamed: 11'] = df2.loc[0, 'Unnamed: 11']
    df2.loc[1, 'Unnamed: 12'] = df2.loc[0, 'Unnamed: 12']
    df2.loc[1, 'Unnamed: 3'] = 'Tên'
    df2.columns = df2.iloc[1]
    df2 = df2.dropna(how='all', ignore_index=True)
    index_council = find_index_hd_in_excel(df2, council_id)
    index_council_next = find_index_hd_in_excel(df2, council_id + 1)
    if index_council_next == -1:
        index_council_next = len(df2)
    df2 = df2.iloc[index_council+2:index_council_next]
    df2 = df2.reset_index(drop=True)
    for i in range(len(df2)):
        df2 = df2.reset_index(drop=True) 
        if df2.loc[i, 'Nhóm'] == group_id:
            students.append(df2.loc[i, 'Mã sinh viên'])
    return students

def find_index_hd_in_excel(df2, id_hd):
    for i in range(len(df2)):
        if df2.loc[i, 'Mã sinh viên'] == f'Hội đồng {id_hd}':
            return i
    return -1
####

#### Process all data
def get_all_councils():
    file_path = os.path.join('DataBase', 'db.xlsx')
    df = pd.read_excel(file_path, sheet_name="DS Hội đồng_DN", skiprows=2)
    df = df.drop(columns = ['Unnamed: 4'])
    df = df.dropna(ignore_index=True)
    df = df[df['Họ và tên'] != 'Họ và tên']
    councils = {}
    for i in range(len(df)):
        member = df.iloc[i]
        name = member['Họ và tên']
        role = member['Nhiệm vụ'].replace('\xa0', '')
        abbr_unit = member['Đơn vị']
        unit = member['Đơn vị.1']
        id_council = f"HD{(i)//5 + 1}"
        if councils.get(id_council) is None:
            councils[id_council] = [{
                'name': name,
                'role': role,
                'abbr_unit': abbr_unit,
                'unit': unit
            }]
        else:
            councils[id_council].append({
                'name': name,
                'role': role,
                'abbr_unit': abbr_unit,
                'unit': unit
            })
    return councils

def get_all_students():
    # file_path = os.path.join('DataBase', 'db.xlsx')
    # df2 = pd.read_excel(file_path, sheet_name="DS SV các Hội đồng phân PB", skiprows=1)
    # df2.loc[1, 'Unnamed: 11'] = df2.loc[0, 'Unnamed: 11']
    # df2.loc[1, 'Unnamed: 12'] = df2.loc[0, 'Unnamed: 12']
    # df2.loc[1, 'Unnamed: 3'] = 'Tên'
    # df2.loc[1, 'Unnamed: 0'] = 'TT'
    # df2.columns = df2.iloc[1]
    # df2 = df2.dropna(how='all', ignore_index=True)
    # councils = get_all_councils()
    # id_council = 'HD1'
    # students = {
    # }
    # for i in range(len(df2)):
    #     msv = df2.loc[i, 'Mã sinh viên']
    #     if 'Hội đồng' in msv:
    #         tmp = msv.split(' ')
    #         id_council = f"HD{tmp[-1]}"
    #         continue
    #     if msv == 'Mã sinh viên':
    #         continue
    #     name = df2.loc[i, 'Họ và tên'] + ' ' + df2.loc[i, 'Tên']
    #     day_of_birth = df2.loc[i, 'Năm sinh']
    #     class_name = df2.loc[i, 'Lớp']
    #     instructor = df2.loc[i, 'Giáo viên hướng dẫn']
    #     subject = df2.loc[i, 'Bộ môn ']
    #     project = df2.loc[i, 'Tên đề tài đồ án/ khóa luận tốt nghiệp']
    #     project_type = df2.loc[i, 'Loại đồ án']
    #     group = id_council + ' - ' + str(df2.loc[i, 'Nhóm']) if not pd.isnull(df2.loc[i, 'Nhóm']) else ''
    #     opponent = df2.loc[i, 'Phản biện  (Khoa)'] if not pd.isna(df2.loc[i, 'Phản biện  (Khoa)']) else ''
    #     council = councils[id_council]
    #     students[msv] = {
    #         'name': name,
    #         'day_of_birth': day_of_birth,
    #         'class_name': class_name,
    #         'instructor': instructor,
    #         'subject': subject,
    #         'project': project,
    #         'project_type': project_type,
    #         'group': group,
    #         'opponent': opponent,
    #         'council': council,
    #         'msv': msv,
    #         'tt': int(tt)
    #     }
    #     # Get students from final sheet
    # new_students = students.copy()
    # for msv in students.keys():
    #     student = students[msv]
    #     if student['project'] == 'None':
    #         continue
    #     project_name = student['project']
    #     students_same_project = get_students_by_project_name(project_name)
    #     for std in students_same_project:
    #         msv_std = std['msv']
    #         if msv_std not in students.keys():
    #             fullname = std['fullname']
    #             day_of_birth = std['day_of_birth']
    #             class_name = std['class']

    #             new_students[msv_std] = students[msv].copy()
    #             new_students[msv_std]['msv'] = msv_std
    #             new_students[msv_std]['name'] = fullname
    #             new_students[msv_std]['day_of_birth'] = day_of_birth
    #             new_students[msv_std]['class_name'] = class_name

    new_students = []
    file_path = os.path.join('DataBase', 'students.json')
    with open(file_path, 'r', encoding='utf-8') as f:
        new_students = json.load(f)
    return new_students
#### End process all data

def index(request):
    lecturers = get_lecturers()
    return render(request, 'index.html', {'lecturers': lecturers})

def hoiDongChuyenMon(request):
    if request.method == 'POST':
        try:
            msv_list = request.GET.getlist('msv', '')
            council_id, group_id = None, None
            group = request.GET.get('group', None)
            data = json.loads(request.body.decode('utf-8'))  # Lấy dữ liệu từ body
            students = data.get('students', [])
            data = data.get('data', {})
            studentsData = [students[msv] for msv in msv_list]
            name = data.get('name', '')
            project_type = data.get('projectType', '')
            projectName = [data.get('projectName', '')]
            unit = data.get('unit', '')
            studentsMsvSameGroup = []
            if group is not None and group != '':
                tmp = group.split(' - ')
                council_id = int(tmp[0][2:])
                group_id = int(tmp[1])
                studentsMsvSameGroup = find_student_by_council_and_group_id(council_id, group_id)
                studentsData = [students[msv] for msv in studentsMsvSameGroup]
                projectName = [f"{studentData['project']} ({studentData['msv']})" for studentData in studentsData]
            for studentData in studentsData:
                studentData['grade'] = getGradeOfStudent(studentData['msv'], name, "hoiDongChuyenMon")
        except (json.JSONDecodeError, AttributeError):
            return JsonResponse({'error': 'Dữ liệu không hợp lệ hoặc trống'}, status=400)

        context = {
            'students': studentsData,
            'students_count': len(studentsData),
            'name': name,
            'project_type': project_type,
            'projects_name': projectName,
            'unit': unit
        }
        return render(request, 'hoiDongChuyenMon.html', context)
    
    # Xử lý GET request (truy cập trực tiếp từ trình duyệt)
    context = {
        'students': [],
        'students_count': 0,
        'name': '',
        'project_type': '',
        'projects_name': [],
        'studentsSameGroup': []
    }
    return render(request, 'hoiDongChuyenMon.html', context)

def baoCaoTienDoL1(request):
    if request.method == 'POST':
        try:
            msv_list = request.GET.getlist('msv', '')
            data = json.loads(request.body.decode('utf-8'))  # Lấy dữ liệu từ body
            data = data.get('data', {})
            students = data.get('students', [])
            students = [students[msv] for msv in msv_list]
            name = data.get('name', '')
            project_type = data.get('projectType', '')
            projectName = data.get('projectName', '')
            for studentData in students:
                studentData['grade'] = getGradeOfStudent(studentData['msv'], name, "baoCaoTienDoL1")
        except (json.JSONDecodeError, AttributeError):
            return JsonResponse({'error': 'Dữ liệu không hợp lệ hoặc trống'}, status=400)

        context = {
            'students': students,
            'students_count': len(students),
            'name': name,
            'project_type': project_type,
            'project_name': projectName
        }
        return render(request, 'baoCaoTienDoL1.html', context)
    
    # Xử lý GET request (truy cập trực tiếp từ trình duyệt)
    context = {
        'students': [],
        'students_count': 0,
        'name': '',
        'project_type': '',
        'project_name': ''
    }
    return render(request, 'baoCaoTienDoL1.html', context)

def baoCaoTienDoL2(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))  # Lấy dữ liệu từ body
            data = data.get('data', {})
            students = data.get('students', [])
            students = list(students.values())
            name = data.get('name', '')
            project_type = data.get('projectType', '')
            projectName = data.get('projectName', '')
            for studentData in students:
                studentData['grade'] = getGradeOfStudent(studentData['msv'], name, "baoCaoTienDoL2")
        except (json.JSONDecodeError, AttributeError):
            return JsonResponse({'error': 'Dữ liệu không hợp lệ hoặc trống'}, status=400)

        context = {
            'students': students,
            'students_count': len(students),
            'name': name,
            'project_type': project_type,
            'project_name': projectName
        }
        return render(request, 'baoCaoTienDoL2.html', context)
    
    # Xử lý GET request (truy cập trực tiếp từ trình duyệt)
    context = {
        'students': [],
        'students_count': 0,
        'name': '',
        'project_type': '',
        'project_name': ''
    }
    return render(request, 'baoCaoTienDoL2.html', context)

def huongdan3(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))  # Lấy dữ liệu từ body
            data = data.get('data', {})
            students = data.get('students', [])
            students = list(students.values())
            name = data.get('name', '')
            project_type = data.get('projectType', '')
            projectName = data.get('projectName', '')
            for studentData in students:
                studentData['grade'] = getGradeOfStudent(studentData['msv'], name, "huongdan3")
        except (json.JSONDecodeError, AttributeError):
            return JsonResponse({'error': 'Dữ liệu không hợp lệ hoặc trống'}, status=400)

        context = {
            'students': students,
            'students_count': len(students),
            'name': name,
            'project_type': project_type,
            'project_name': projectName
        }
        return render(request, 'huongdan3.html', context)
    
    # Xử lý GET request (truy cập trực tiếp từ trình duyệt)
    context = {
        'students': [],
        'students_count': 0,
        'name': '',
        'project_type': '',
        'project_name': ''
    }
    return render(request, 'huongdan3.html', context)

def canBoPhanBien(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))  # Lấy dữ liệu từ body
            data = data.get('data', {})
            students = data.get('students', [])
            students = list(students.values())
            name = data.get('name', '')
            project_type = data.get('projectType', '')
            projectName = data.get('projectName', '')
            for studentData in students:
                studentData['grade'] = getGradeOfStudent(studentData['msv'], name, "canBoPhanBien")
        except (json.JSONDecodeError, AttributeError):
            return JsonResponse({'error': 'Dữ liệu không hợp lệ hoặc trống'}, status=400)

        context = {
            'students': students,
            'students_count': len(students),
            'name': name,
            'project_type': project_type,
            'project_name': projectName
        }
        return render(request, 'canBoPhanBien.html', context)
    
    # Xử lý GET request (truy cập trực tiếp từ trình duyệt)
    context = {
        'students': [],
        'students_count': 0,
        'name': '',
        'project_type': '',
        'project_name': ''
    }
    return render(request, 'canBoPhanBien.html', context)


# Process submit form
def process_form_hdcm_new(request):
    if request.method == 'POST':
        # Lấy số lượng sinh viên
        try:
            students_count = int(request.POST.get('students_count', 0))
        except ValueError:
            students_count = 0

        # Lấy dữ liệu từ form
        students = []
        for i in range(1, students_count + 1):
            student = {
                'fullname': request.POST.get(f'student_fullname_{i}', '').strip(),
                'msv': request.POST.get(f'student_msv_{i}', '').strip(),
                'class': request.POST.get(f'student_class_{i}', '').strip(),
                'diemC33': request.POST.get(f'diemC33SV{i}', '').strip(),
                'diemC42': request.POST.get(f'diemC42SV{i}', '').strip(),
                'diemC53': request.POST.get(f'diemC53SV{i}', '').strip(),
                'diemC63': request.POST.get(f'diemC63SV{i}', '').strip(),
                'diemC64': request.POST.get(f'diemC64SV{i}', '').strip(),
                'gpa': request.POST.get(f'gpaSV{i}', '').strip()
            }
            if student['msv']:  # Đảm bảo có mã sinh viên
                students.append(student)

        nhanXet = request.POST.get('nhanXet', '').strip()
        lecturer_name = request.POST.get('lecturer_name', '').strip()
        project_type = request.POST.get('project_type', '').strip()
        project_name = request.POST.get('project_name', '').strip()
        form_type = 'hdcm'

        # Đường dẫn tới thư mục và file Excel
        data_dir = 'DataCollected'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)

        file_path = os.path.join(data_dir, 'final_new.xlsx')

        # Mở hoặc tạo mới file Excel
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
            # Lấy header hiện tại
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
            # Thêm các cột mới theo cấu trúc mới
            required_headers = ['Họ và tên', 'Mã sinh viên', 'Lớp', 
            'HDCM_uv1-họ tên', 'HDCM_uv1_C3.3', 'HDCM_uv1_C4.2', 'HDCM_uv1_C5.3', 'HDCM_uv1_C6.3', 'HDCM_uv1_C6.4', 'HDCM_uv1_gpa', 
            'HDCM_uv2-họ tên', 'HDCM_uv2_C3.3', 'HDCM_uv2_C4.2', 'HDCM_uv2_C5.3', 'HDCM_uv2_C6.3', 'HDCM_uv2_C6.4', 'HDCM_uv2_gpa', 
            'HDCM_uv3-họ tên', 'HDCM_uv3_C3.3', 'HDCM_uv3_C4.2', 'HDCM_uv3_C5.3', 'HDCM_uv3_C6.3', 'HDCM_uv3_C6.4', 'HDCM_uv3_gpa', 
            'HDCM_uv4-họ tên', 'HDCM_uv4_C3.3', 'HDCM_uv4_C4.2', 'HDCM_uv4_C5.3', 'HDCM_uv4_C6.3', 'HDCM_uv4_C6.4', 'HDCM_uv4_gpa', 
            'HDCM_uv5-họ tên', 'HDCM_uv5_C3.3', 'HDCM_uv5_C4.2', 'HDCM_uv5_C5.3', 'HDCM_uv5_C6.3', 'HDCM_uv5_C6.4', 'HDCM_uv5_gpa', 
            'CBHD_1-họ tên', 'CBHD_1_C1.1', 'CBHD_1_C1.2', 'CBHD_1_C5.1', 'CBHD_1_gpa', 
            'CBHD_2-họ tên', 'CBHD_2_C2.1', 'CBHD_2_C2.2', 'CBHD_2_C3.1', 'CBHD_2_C5.2', 'CBHD_2_gpa', 
            'CBHD_3-họ tên', 'CBHD_3_C2.3', 'CBHD_3_C3.2', 'CBHD_3_C4.1', 'CBHD_3_C6.1', 'CBHD_3_C6.2', 'CBHD_3_gpa', 
            'CBPB-họ tên', 'CBPB_C2.3', 'CBPB_C3.2', 'CBPB_C4.1', 'CBPB_C6.1', 'CBPB_C6.2', 'CBPB_gpa'
            ]
            # Kiểm tra nếu header hiện tại không khớp với required_headers
            if not header_row or list(header_row)  != required_headers:
                # Xóa header hiện tại
                sheet.delete_rows(1)
                # Thêm header mới vào hàng đầu tiên
                for col_num, header in enumerate(required_headers, 1):
                    sheet.cell(row=1, column=col_num, value=header)

            # Lưu lại workbook
            workbook.save(file_path)   
        sheet = workbook.active

        # Mapping headers to column indexes
        headers = {}
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        for idx, cell in enumerate(header_row):
            if cell:
                headers[cell.strip().lower()] = idx + 1  # 1-based index

        # Lưu lại workbook
        workbook.save(file_path)
        

        # Duyệt qua từng sinh viên và lưu thông tin vào file Excel mới
        for student in students:
            msv = student['msv']
            # Tìm sinh viên theo msv ở cột 'mã sinh viên'
            student_found = False
            for row in sheet.iter_rows(min_row=2, max_col=len(headers), values_only=False):
                cell_msv = row[headers['mã sinh viên'] -1]
                if cell_msv.value == msv:
                    student_found = True
                    # Tìm cột HDCM_uv1 đến HDCM_uv5 trống
                    for uv in range(1, 6):
                        col_name = f"hdcm_uv{uv}-họ tên"
                        if col_name not in headers:
                            continue  # Bỏ qua nếu thiếu cột
                        cell_uv = row[headers[col_name] -1]
                        if not cell_uv.value:
                            # Điền thông tin vào các cột tương ứng
                            cell_uv.value = lecturer_name
                            row[headers[f"hdcm_uv{uv}_c3.3"] -1].value = student['diemC33']
                            row[headers[f"hdcm_uv{uv}_c4.2"] -1].value = student['diemC42']
                            row[headers[f"hdcm_uv{uv}_c5.3"] -1].value = student['diemC53']
                            row[headers[f"hdcm_uv{uv}_c6.3"] -1].value = student['diemC63']
                            row[headers[f"hdcm_uv{uv}_c6.4"] -1].value = student['diemC64']
                            row[headers[f"hdcm_uv{uv}_gpa"] -1].value = student['gpa']
                            print(f"Ghi dữ liệu vào HDCM_uv{uv} cho MSV: {msv}")
                            break  # Ghi xong vào cột trống, dừng tìm cột tiếp theo
                    break  # Đã tìm thấy sinh viên, dừng tìm kiếm

            if not student_found:
                # Thêm sinh viên mới và ghi vào HDCM_uv1
                new_row = [
                    student['fullname'],  # Họ và tên
                    student['msv'],       # Mã sinh viên
                    student['class'],     # Lớp
                    lecturer_name,        # HDCM_uv1-họ tên
                    student['diemC33'],   # HDCM_uv1_C3.3
                    student['diemC42'],   # HDCM_uv1_C4.2
                    student['diemC53'],   # HDCM_uv1_C5.3
                    student['diemC63'],   # HDCM_uv1_C6.3
                    student['diemC64'],   # HDCM_uv1_C6.4
                    student['gpa'],
                    ] + [""] * 53  # Thêm các cột còn lại dưới dạng chuỗi rỗng
                sheet.append(new_row)
        # Lưu file Excel
        workbook.save(file_path)
        xuat()
        # Chuyển hướng đến trang testOutput.html với dữ liệu
        return JsonResponse({'message': 'Chấm điểm thành công'})

    return JsonResponse({'message': 'Chấm điểm không thành công'})

def process_form_hd1_new(request):
    if request.method == 'POST':
        # Lấy số lượng sinh viên
        try:
            students_count = int(request.POST.get('students_count', 0))
        except ValueError:
            students_count = 0

        # Lấy dữ liệu từ form
        students = []
        for i in range(1, students_count + 1):
            student = {
                'fullname': request.POST.get(f'student_fullname_{i}', '').strip(),
                'msv': request.POST.get(f'student_msv_{i}', '').strip(),
                'class': request.POST.get(f'student_class_{i}', '').strip(),
                'diemC11': request.POST.get(f'diemC11SV{i}', '').strip(),
                'diemC12': request.POST.get(f'diemC12SV{i}', '').strip(),
                'diemC51': request.POST.get(f'diemC51SV{i}', '').strip(),
                'gpa': request.POST.get(f'gpaSV{i}', '').strip()
            }
            if student['msv']:
                students.append(student)

        nhanXet = request.POST.get('nhanXet', '').strip()
        lecturer_name = request.POST.get('lecturer_name', '').strip()
        project_type = request.POST.get('project_type', '').strip()
        project_name = request.POST.get('project_name', '').strip()
        form_type = 'hd1_new'

        # Đường dẫn tới thư mục và file Excel
        data_dir = 'DataCollected'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)

        file_path = os.path.join(data_dir, 'final_new.xlsx')

        # Mở hoặc tạo mới file Excel
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
            # Lấy header hiện tại
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
            # Thêm các cột mới theo cấu trúc mới
            required_headers = ['Họ và tên', 'Mã sinh viên', 'Lớp', 
            'HDCM_uv1-họ tên', 'HDCM_uv1_C3.3', 'HDCM_uv1_C4.2', 'HDCM_uv1_C5.3', 'HDCM_uv1_C6.3', 'HDCM_uv1_C6.4', 'HDCM_uv1_gpa', 
            'HDCM_uv2-họ tên', 'HDCM_uv2_C3.3', 'HDCM_uv2_C4.2', 'HDCM_uv2_C5.3', 'HDCM_uv2_C6.3', 'HDCM_uv2_C6.4', 'HDCM_uv2_gpa', 
            'HDCM_uv3-họ tên', 'HDCM_uv3_C3.3', 'HDCM_uv3_C4.2', 'HDCM_uv3_C5.3', 'HDCM_uv3_C6.3', 'HDCM_uv3_C6.4', 'HDCM_uv3_gpa', 
            'HDCM_uv4-họ tên', 'HDCM_uv4_C3.3', 'HDCM_uv4_C4.2', 'HDCM_uv4_C5.3', 'HDCM_uv4_C6.3', 'HDCM_uv4_C6.4', 'HDCM_uv4_gpa', 
            'HDCM_uv5-họ tên', 'HDCM_uv5_C3.3', 'HDCM_uv5_C4.2', 'HDCM_uv5_C5.3', 'HDCM_uv5_C6.3', 'HDCM_uv5_C6.4', 'HDCM_uv5_gpa', 
            'CBHD_1-họ tên', 'CBHD_1_C1.1', 'CBHD_1_C1.2', 'CBHD_1_C5.1', 'CBHD_1_gpa', 
            'CBHD_2-họ tên', 'CBHD_2_C2.1', 'CBHD_2_C2.2', 'CBHD_2_C3.1', 'CBHD_2_C5.2', 'CBHD_2_gpa', 
            'CBHD_3-họ tên', 'CBHD_3_C2.3', 'CBHD_3_C3.2', 'CBHD_3_C4.1', 'CBHD_3_C6.1', 'CBHD_3_C6.2', 'CBHD_3_gpa', 
            'CBPB-họ tên', 'CBPB_C2.3', 'CBPB_C3.2', 'CBPB_C4.1', 'CBPB_C6.1', 'CBPB_C6.2', 'CBPB_gpa'
            ]
            # Kiểm tra nếu header hiện tại không khớp với required_headers
            if not header_row or list(header_row)  != required_headers:
                # Xóa header hiện tại
                sheet.delete_rows(1)
                # Thêm header mới vào hàng đầu tiên
                for col_num, header in enumerate(required_headers, 1):
                    sheet.cell(row=1, column=col_num, value=header)

            # Lưu lại workbook
            workbook.save(file_path)   
        sheet = workbook.active
            
        

        # Ánh xạ header => index cột
        headers = {}
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        for idx, cell in enumerate(header_row):
            if cell:
                headers[cell.strip().lower()] = idx + 1

        # Kiểm tra nếu header hiện tại không khớp với required_headers
        if list(headers.keys()) != required_headers:
            # Xóa header hiện tại
            sheet.delete_rows(1)
            # Chèn hàng mới ở vị trí đầu tiên
            sheet.insert_rows(1)
            # Thêm header mới vào hàng đầu tiên
            for col_num, header in enumerate(required_headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

        # Lưu lại workbook
        workbook.save(file_path)
        
        

        # Duyệt qua từng sinh viên
        for student in students:
            msv = student['msv']
            student_found = False
            for row in sheet.iter_rows(min_row=2, max_col=len(headers), values_only=False):
                cell_msv = row[headers['mã sinh viên'] - 1]
                # Nếu đã có msv, kiểm tra cột CBHD_1-họ tên trống để ghi
                if cell_msv.value == msv:
                    student_found = True
                    cell_cbhd_name = row[headers['cbhd_1-họ tên'] - 1]
                    if not cell_cbhd_name.value:  
                        cell_cbhd_name.value = lecturer_name
                        row[headers['cbhd_1_c1.1'] - 1].value = student['diemC11']
                        row[headers['cbhd_1_c1.2'] - 1].value = student['diemC12']
                        row[headers['cbhd_1_c5.1'] - 1].value = student['diemC51']
                        row[headers['cbhd_1_gpa'] - 1].value = student['gpa']
                    break

            # Nếu chưa có msv trong file
            if not student_found:
                # Tạo dòng mới
                new_row = [
                    student['fullname'],        
                    student['msv'],            
                    student['class'],             
                ] +[""]*35 + [
                    lecturer_name ,
                    student['diemC11'],         
                    student['diemC12'],         
                    student['diemC51'],
                    student['gpa']

                ] + [""]*20
                sheet.append(new_row)

        # Lưu file Excel
        workbook.save(file_path)
        xuat()
        # Chuyển hướng đến trang testOutput.html
        return JsonResponse({'message': 'Chấm điểm thành công'})

    # Nếu không phải POST, chuyển về baoCaoTienDoL1
    return JsonResponse({'message': 'Chấm điểm không thành công'})

def process_form_hd2_new(request):
    if request.method == 'POST':
        # Lấy số lượng sinh viên
        try:
            students_count = int(request.POST.get('students_count', 0))
        except ValueError:
            students_count = 0

        # Lấy dữ liệu từ form
        students = []
        for i in range(1, students_count + 1):
            student = {
                'fullname': request.POST.get(f'student_fullname_{i}', '').strip(),
                'msv': request.POST.get(f'student_msv_{i}', '').strip(),
                'class': request.POST.get(f'student_class_{i}', '').strip(),
                'diemC21': request.POST.get(f'diemC21SV{i}', '').strip(),
                'diemC22': request.POST.get(f'diemC22SV{i}', '').strip(),
                'diemC31': request.POST.get(f'diemC31SV{i}', '').strip(),
                'diemC52': request.POST.get(f'diemC52SV{i}', '').strip(),
                'gpa': request.POST.get(f'gpaSV{i}', '').strip()
            }
            if student['msv']:
                students.append(student)

        nhanXet = request.POST.get('nhanXet', '').strip()
        lecturer_name = request.POST.get('lecturer_name', '').strip()
        project_type = request.POST.get('project_type', '').strip()
        project_name = request.POST.get('project_name', '').strip()
        form_type = 'hd1_new'

        # Đường dẫn tới thư mục và file Excel
        data_dir = 'DataCollected'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)

        file_path = os.path.join(data_dir, 'final_new.xlsx')

        # Mở hoặc tạo mới file Excel
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
            # Lấy header hiện tại
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
            # Thêm các cột mới theo cấu trúc mới
            required_headers = ['Họ và tên', 'Mã sinh viên', 'Lớp', 
            'HDCM_uv1-họ tên', 'HDCM_uv1_C3.3', 'HDCM_uv1_C4.2', 'HDCM_uv1_C5.3', 'HDCM_uv1_C6.3', 'HDCM_uv1_C6.4', 'HDCM_uv1_gpa', 
            'HDCM_uv2-họ tên', 'HDCM_uv2_C3.3', 'HDCM_uv2_C4.2', 'HDCM_uv2_C5.3', 'HDCM_uv2_C6.3', 'HDCM_uv2_C6.4', 'HDCM_uv2_gpa', 
            'HDCM_uv3-họ tên', 'HDCM_uv3_C3.3', 'HDCM_uv3_C4.2', 'HDCM_uv3_C5.3', 'HDCM_uv3_C6.3', 'HDCM_uv3_C6.4', 'HDCM_uv3_gpa', 
            'HDCM_uv4-họ tên', 'HDCM_uv4_C3.3', 'HDCM_uv4_C4.2', 'HDCM_uv4_C5.3', 'HDCM_uv4_C6.3', 'HDCM_uv4_C6.4', 'HDCM_uv4_gpa', 
            'HDCM_uv5-họ tên', 'HDCM_uv5_C3.3', 'HDCM_uv5_C4.2', 'HDCM_uv5_C5.3', 'HDCM_uv5_C6.3', 'HDCM_uv5_C6.4', 'HDCM_uv5_gpa', 
            'CBHD_1-họ tên', 'CBHD_1_C1.1', 'CBHD_1_C1.2', 'CBHD_1_C5.1', 'CBHD_1_gpa', 
            'CBHD_2-họ tên', 'CBHD_2_C2.1', 'CBHD_2_C2.2', 'CBHD_2_C3.1', 'CBHD_2_C5.2', 'CBHD_2_gpa', 
            'CBHD_3-họ tên', 'CBHD_3_C2.3', 'CBHD_3_C3.2', 'CBHD_3_C4.1', 'CBHD_3_C6.1', 'CBHD_3_C6.2', 'CBHD_3_gpa', 
            'CBPB-họ tên', 'CBPB_C2.3', 'CBPB_C3.2', 'CBPB_C4.1', 'CBPB_C6.1', 'CBPB_C6.2', 'CBPB_gpa'
            ]
            # Kiểm tra nếu header hiện tại không khớp với required_headers
            if not header_row or list(header_row)  != required_headers:
                # Xóa header hiện tại
                sheet.delete_rows(1)
                # Thêm header mới vào hàng đầu tiên
                for col_num, header in enumerate(required_headers, 1):
                    sheet.cell(row=1, column=col_num, value=header)

            # Lưu lại workbook
            workbook.save(file_path)   
        sheet = workbook.active
            
        

        # Ánh xạ header => index cột
        headers = {}
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        for idx, cell in enumerate(header_row):
            if cell:
                headers[cell.strip().lower()] = idx + 1

        required_headers = ['Họ và tên', 'Mã sinh viên', 'Lớp', 
                            'HDCM_uv1-họ tên', 'HDCM_uv1_C3.3', 'HDCM_uv1_C4.2', 'HDCM_uv1_C5.3', 'HDCM_uv1_C6.3', 'HDCM_uv1_C6.4', 'HDCM_uv1_gpa', 
                            'HDCM_uv2-họ tên', 'HDCM_uv2_C3.3', 'HDCM_uv2_C4.2', 'HDCM_uv2_C5.3', 'HDCM_uv2_C6.3', 'HDCM_uv2_C6.4', 'HDCM_uv2_gpa', 
                            'HDCM_uv3-họ tên', 'HDCM_uv3_C3.3', 'HDCM_uv3_C4.2', 'HDCM_uv3_C5.3', 'HDCM_uv3_C6.3', 'HDCM_uv3_C6.4', 'HDCM_uv3_gpa', 
                            'HDCM_uv4-họ tên', 'HDCM_uv4_C3.3', 'HDCM_uv4_C4.2', 'HDCM_uv4_C5.3', 'HDCM_uv4_C6.3', 'HDCM_uv4_C6.4', 'HDCM_uv4_gpa', 
                            'HDCM_uv5-họ tên', 'HDCM_uv5_C3.3', 'HDCM_uv5_C4.2', 'HDCM_uv5_C5.3', 'HDCM_uv5_C6.3', 'HDCM_uv5_C6.4', 'HDCM_uv5_gpa', 
                            'CBHD_1-họ tên', 'CBHD_1_C1.1', 'CBHD_1_C1.2', 'CBHD_1_C5.1', 'CBHD_1_gpa', 
                            'CBHD_2-họ tên', 'CBHD_2_C2.1', 'CBHD_2_C2.2', 'CBHD_2_C3.1', 'CBHD_2_C5.2', 'CBHD_2_gpa', 
                            'CBHD_3-họ tên', 'CBHD_3_C2.3', 'CBHD_3_C3.2', 'CBHD_3_C4.1', 'CBHD_3_C6.1', 'CBHD_3_C6.2', 'CBHD_3_gpa', 
                            'CBPB-họ tên', 'CBPB_C2.3', 'CBPB_C3.2', 'CBPB_C4.1', 'CBPB_C6.1', 'CBPB_C6.2', 'CBPB_gpa'
                            ]

        # Duyệt qua từng sinh viên
        for student in students:
            msv = student['msv']
            student_found = False
            for row in sheet.iter_rows(min_row=2, max_col=len(headers), values_only=False):
                cell_msv = row[headers['mã sinh viên'] - 1]
                # Nếu đã có msv, kiểm tra cột CBHD_1-họ tên trống để ghi
                if cell_msv.value == msv:
                    student_found = True
                    cell_cbhd_name = row[headers['cbhd_2-họ tên'] - 1]
                    if not cell_cbhd_name.value:  
                        cell_cbhd_name.value = lecturer_name
                        row[headers['cbhd_2_c2.1'] - 1].value = student['diemC21']
                        row[headers['cbhd_2_c2.2'] - 1].value = student['diemC22']
                        row[headers['cbhd_2_c3.1'] - 1].value = student['diemC31']
                        row[headers['cbhd_2_c5.2'] - 1].value = student['diemC52']
                        row[headers['cbhd_2_gpa'] - 1].value = student['gpa']

                    break

            # Nếu chưa có msv trong file
            if not student_found:
                # Tạo dòng mới
                new_row = [
                    student['fullname'],        
                    student['msv'],            
                    student['class'],             
                ] +[""]*40 + [
                    lecturer_name ,
                    student['diemC21'],         
                    student['diemC22'],         
                    student['diemC31'],
                    student['diemC52'],
                    student['gpa'],

                ] + [""]*14
                sheet.append(new_row)

        # Lưu file Excel
        workbook.save(file_path)
        xuat()
        # Chuyển hướng đến trang testOutput.html
        return JsonResponse({'message': 'Chấm điểm thành công'})

    # Nếu không phải POST, chuyển về baoCaoTienDoL1
    return JsonResponse({'message': 'Chấm điểm không thành công'})


def process_form_hd3_new(request):
    if request.method == 'POST':
        # Lấy số lượng sinh viên
        try:
            students_count = int(request.POST.get('students_count', 0))
        except ValueError:
            students_count = 0

        # Lấy dữ liệu từ form
        students = []
        for i in range(1, students_count + 1):
            student = {
                'fullname': request.POST.get(f'student_fullname_{i}', '').strip(),
                'msv': request.POST.get(f'student_msv_{i}', '').strip(),
                'class': request.POST.get(f'student_class_{i}', '').strip(),
                'diemC23': request.POST.get(f'diemC23SV{i}', '').strip(),
                'diemC32': request.POST.get(f'diemC32SV{i}', '').strip(),
                'diemC41': request.POST.get(f'diemC41SV{i}', '').strip(),
                'diemC61': request.POST.get(f'diemC61SV{i}', '').strip(),
                'diemC62': request.POST.get(f'diemC62SV{i}', '').strip(),
                'gpa': request.POST.get(f'gpaSV{i}', '').strip()
            }
            if student['msv']:
                students.append(student)

        nhanXet = request.POST.get('nhanXet', '').strip()
        lecturer_name = request.POST.get('lecturer_name', '').strip()
        project_type = request.POST.get('project_type', '').strip()
        project_name = request.POST.get('project_name', '').strip()
        form_type = 'hd1_new'

        # Đường dẫn tới thư mục và file Excel
        data_dir = 'DataCollected'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)

        file_path = os.path.join(data_dir, 'final_new.xlsx')

        # Mở hoặc tạo mới file Excel
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
            # Lấy header hiện tại
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
            # Thêm các cột mới theo cấu trúc mới
            required_headers = ['Họ và tên', 'Mã sinh viên', 'Lớp', 
            'HDCM_uv1-họ tên', 'HDCM_uv1_C3.3', 'HDCM_uv1_C4.2', 'HDCM_uv1_C5.3', 'HDCM_uv1_C6.3', 'HDCM_uv1_C6.4', 'HDCM_uv1_gpa', 
            'HDCM_uv2-họ tên', 'HDCM_uv2_C3.3', 'HDCM_uv2_C4.2', 'HDCM_uv2_C5.3', 'HDCM_uv2_C6.3', 'HDCM_uv2_C6.4', 'HDCM_uv2_gpa', 
            'HDCM_uv3-họ tên', 'HDCM_uv3_C3.3', 'HDCM_uv3_C4.2', 'HDCM_uv3_C5.3', 'HDCM_uv3_C6.3', 'HDCM_uv3_C6.4', 'HDCM_uv3_gpa', 
            'HDCM_uv4-họ tên', 'HDCM_uv4_C3.3', 'HDCM_uv4_C4.2', 'HDCM_uv4_C5.3', 'HDCM_uv4_C6.3', 'HDCM_uv4_C6.4', 'HDCM_uv4_gpa', 
            'HDCM_uv5-họ tên', 'HDCM_uv5_C3.3', 'HDCM_uv5_C4.2', 'HDCM_uv5_C5.3', 'HDCM_uv5_C6.3', 'HDCM_uv5_C6.4', 'HDCM_uv5_gpa', 
            'CBHD_1-họ tên', 'CBHD_1_C1.1', 'CBHD_1_C1.2', 'CBHD_1_C5.1', 'CBHD_1_gpa', 
            'CBHD_2-họ tên', 'CBHD_2_C2.1', 'CBHD_2_C2.2', 'CBHD_2_C3.1', 'CBHD_2_C5.2', 'CBHD_2_gpa', 
            'CBHD_3-họ tên', 'CBHD_3_C2.3', 'CBHD_3_C3.2', 'CBHD_3_C4.1', 'CBHD_3_C6.1', 'CBHD_3_C6.2', 'CBHD_3_gpa', 
            'CBPB-họ tên', 'CBPB_C2.3', 'CBPB_C3.2', 'CBPB_C4.1', 'CBPB_C6.1', 'CBPB_C6.2', 'CBPB_gpa'
            ]
            # Kiểm tra nếu header hiện tại không khớp với required_headers
            if not header_row or list(header_row)  != required_headers:
                # Xóa header hiện tại
                sheet.delete_rows(1)
                # Thêm header mới vào hàng đầu tiên
                for col_num, header in enumerate(required_headers, 1):
                    sheet.cell(row=1, column=col_num, value=header)

            # Lưu lại workbook
            workbook.save(file_path)   
        sheet = workbook.active

        # Ánh xạ header => index cột
        headers = {}
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        for idx, cell in enumerate(header_row):
            if cell:
                headers[cell.strip().lower()] = idx + 1

        required_headers = ['Họ và tên', 'Mã sinh viên', 'Lớp', 
                            'HDCM_uv1-họ tên', 'HDCM_uv1_C3.3', 'HDCM_uv1_C4.2', 'HDCM_uv1_C5.3', 'HDCM_uv1_C6.3', 'HDCM_uv1_C6.4', 'HDCM_uv1_gpa', 
                            'HDCM_uv2-họ tên', 'HDCM_uv2_C3.3', 'HDCM_uv2_C4.2', 'HDCM_uv2_C5.3', 'HDCM_uv2_C6.3', 'HDCM_uv2_C6.4', 'HDCM_uv2_gpa', 
                            'HDCM_uv3-họ tên', 'HDCM_uv3_C3.3', 'HDCM_uv3_C4.2', 'HDCM_uv3_C5.3', 'HDCM_uv3_C6.3', 'HDCM_uv3_C6.4', 'HDCM_uv3_gpa', 
                            'HDCM_uv4-họ tên', 'HDCM_uv4_C3.3', 'HDCM_uv4_C4.2', 'HDCM_uv4_C5.3', 'HDCM_uv4_C6.3', 'HDCM_uv4_C6.4', 'HDCM_uv4_gpa', 
                            'HDCM_uv5-họ tên', 'HDCM_uv5_C3.3', 'HDCM_uv5_C4.2', 'HDCM_uv5_C5.3', 'HDCM_uv5_C6.3', 'HDCM_uv5_C6.4', 'HDCM_uv5_gpa', 
                            'CBHD_1-họ tên', 'CBHD_1_C1.1', 'CBHD_1_C1.2', 'CBHD_1_C5.1', 'CBHD_1_gpa', 
                            'CBHD_2-họ tên', 'CBHD_2_C2.1', 'CBHD_2_C2.2', 'CBHD_2_C3.1', 'CBHD_2_C5.2', 'CBHD_2_gpa', 
                            'CBHD_3-họ tên', 'CBHD_3_C2.3', 'CBHD_3_C3.2', 'CBHD_3_C4.1', 'CBHD_3_C6.1', 'CBHD_3_C6.2', 'CBHD_3_gpa', 
                            'CBPB-họ tên', 'CBPB_C2.3', 'CBPB_C3.2', 'CBPB_C4.1', 'CBPB_C6.1', 'CBPB_C6.2', 'CBPB_gpa'
                            ]
        

        # Duyệt qua từng sinh viên
        for student in students:
            msv = student['msv']
            student_found = False
            for row in sheet.iter_rows(min_row=2, max_col=len(headers), values_only=False):
                cell_msv = row[headers['mã sinh viên'] - 1]
                # Nếu đã có msv, kiểm tra cột CBHD_1-họ tên trống để ghi
                if cell_msv.value == msv:
                    student_found = True
                    cell_cbhd_name = row[headers['cbhd_3-họ tên'] - 1]
                    if not cell_cbhd_name.value:  
                        cell_cbhd_name.value = lecturer_name
                        row[headers['cbhd_3_c2.3'] - 1].value = student['diemC23']
                        row[headers['cbhd_3_c3.2'] - 1].value = student['diemC32']
                        row[headers['cbhd_3_c4.1'] - 1].value = student['diemC41']
                        row[headers['cbhd_3_c6.1'] - 1].value = student['diemC61']
                        row[headers['cbhd_3_c6.2'] - 1].value = student['diemC62']
                        row[headers['cbhd_3_gpa'] - 1].value = student['gpa']
                    break

            # Nếu chưa có msv trong file
            if not student_found:
                # Tạo dòng mới
                new_row = [
                    student['fullname'],        
                    student['msv'],            
                    student['class'],             
                ] +[""]*46 + [
                    lecturer_name ,
                    student['diemC23'],         
                    student['diemC32'],         
                    student['diemC41'],
                    student['diemC61'],
                    student['diemC62'],
                    student['gpa'],
                ] + [""]*7
                sheet.append(new_row)

        # Lưu file Excel
        workbook.save(file_path)
        xuat()
        return JsonResponse({'message': 'Chấm điểm thành công'})

    # Nếu không phải POST, chuyển về baoCaoTienDoL1
    return JsonResponse({'message': 'Chấm điểm không thành công'})
                    
def process_form_pb_new(request):
    if request.method == 'POST':
        # Lấy số lượng sinh viên
        try:
            students_count = int(request.POST.get('students_count', 0))
        except ValueError:
            students_count = 0

        # Lấy dữ liệu từ form
        students = []
        for i in range(1, students_count + 1):
            student = {
                'fullname': request.POST.get(f'student_fullname_{i}', '').strip(),
                'msv': request.POST.get(f'student_msv_{i}', '').strip(),
                'class': request.POST.get(f'student_class_{i}', '').strip(),
                'diemC23': request.POST.get(f'diemC23SV{i}', '').strip(),
                'diemC32': request.POST.get(f'diemC32SV{i}', '').strip(),
                'diemC41': request.POST.get(f'diemC41SV{i}', '').strip(),
                'diemC61': request.POST.get(f'diemC61SV{i}', '').strip(),
                'diemC62': request.POST.get(f'diemC62SV{i}', '').strip(),
                'gpa': request.POST.get(f'gpaSV{i}', '').strip()
            }
            if student['msv']:
                students.append(student)

        nhanXet = request.POST.get('nhanXet', '').strip()
        lecturer_name = request.POST.get('lecturer_name', '').strip()
        project_type = request.POST.get('project_type', '').strip()
        project_name = request.POST.get('project_name', '').strip()
        form_type = 'hd1_new'

        # Đường dẫn tới thư mục và file Excel
        data_dir = 'DataCollected'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)

        file_path = os.path.join(data_dir, 'final_new.xlsx')

        # Mở hoặc tạo mới file Excel
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
            # Lấy header hiện tại
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
            # Thêm các cột mới theo cấu trúc mới
            required_headers = ['Họ và tên', 'Mã sinh viên', 'Lớp', 
            'HDCM_uv1-họ tên', 'HDCM_uv1_C3.3', 'HDCM_uv1_C4.2', 'HDCM_uv1_C5.3', 'HDCM_uv1_C6.3', 'HDCM_uv1_C6.4', 'HDCM_uv1_gpa', 
            'HDCM_uv2-họ tên', 'HDCM_uv2_C3.3', 'HDCM_uv2_C4.2', 'HDCM_uv2_C5.3', 'HDCM_uv2_C6.3', 'HDCM_uv2_C6.4', 'HDCM_uv2_gpa', 
            'HDCM_uv3-họ tên', 'HDCM_uv3_C3.3', 'HDCM_uv3_C4.2', 'HDCM_uv3_C5.3', 'HDCM_uv3_C6.3', 'HDCM_uv3_C6.4', 'HDCM_uv3_gpa', 
            'HDCM_uv4-họ tên', 'HDCM_uv4_C3.3', 'HDCM_uv4_C4.2', 'HDCM_uv4_C5.3', 'HDCM_uv4_C6.3', 'HDCM_uv4_C6.4', 'HDCM_uv4_gpa', 
            'HDCM_uv5-họ tên', 'HDCM_uv5_C3.3', 'HDCM_uv5_C4.2', 'HDCM_uv5_C5.3', 'HDCM_uv5_C6.3', 'HDCM_uv5_C6.4', 'HDCM_uv5_gpa', 
            'CBHD_1-họ tên', 'CBHD_1_C1.1', 'CBHD_1_C1.2', 'CBHD_1_C5.1', 'CBHD_1_gpa', 
            'CBHD_2-họ tên', 'CBHD_2_C2.1', 'CBHD_2_C2.2', 'CBHD_2_C3.1', 'CBHD_2_C5.2', 'CBHD_2_gpa', 
            'CBHD_3-họ tên', 'CBHD_3_C2.3', 'CBHD_3_C3.2', 'CBHD_3_C4.1', 'CBHD_3_C6.1', 'CBHD_3_C6.2', 'CBHD_3_gpa', 
            'CBPB-họ tên', 'CBPB_C2.3', 'CBPB_C3.2', 'CBPB_C4.1', 'CBPB_C6.1', 'CBPB_C6.2', 'CBPB_gpa'
            ]
            # Kiểm tra nếu header hiện tại không khớp với required_headers
            if not header_row or list(header_row)  != required_headers:
                # Xóa header hiện tại
                sheet.delete_rows(1)
                # Thêm header mới vào hàng đầu tiên
                for col_num, header in enumerate(required_headers, 1):
                    sheet.cell(row=1, column=col_num, value=header)

            # Lưu lại workbook
            workbook.save(file_path)  
        sheet = workbook.active         

        # Ánh xạ header => index cột
        headers = {}
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        for idx, cell in enumerate(header_row):
            if cell:
                headers[cell.strip().lower()] = idx + 1

        required_headers = ['Họ và tên', 'Mã sinh viên', 'Lớp', 
            'HDCM_uv1-họ tên', 'HDCM_uv1_C3.3', 'HDCM_uv1_C4.2', 'HDCM_uv1_C5.3', 'HDCM_uv1_C6.3', 'HDCM_uv1_C6.4', 'HDCM_uv1_gpa', 
            'HDCM_uv2-họ tên', 'HDCM_uv2_C3.3', 'HDCM_uv2_C4.2', 'HDCM_uv2_C5.3', 'HDCM_uv2_C6.3', 'HDCM_uv2_C6.4', 'HDCM_uv2_gpa', 
            'HDCM_uv3-họ tên', 'HDCM_uv3_C3.3', 'HDCM_uv3_C4.2', 'HDCM_uv3_C5.3', 'HDCM_uv3_C6.3', 'HDCM_uv3_C6.4', 'HDCM_uv3_gpa', 
            'HDCM_uv4-họ tên', 'HDCM_uv4_C3.3', 'HDCM_uv4_C4.2', 'HDCM_uv4_C5.3', 'HDCM_uv4_C6.3', 'HDCM_uv4_C6.4', 'HDCM_uv4_gpa', 
            'HDCM_uv5-họ tên', 'HDCM_uv5_C3.3', 'HDCM_uv5_C4.2', 'HDCM_uv5_C5.3', 'HDCM_uv5_C6.3', 'HDCM_uv5_C6.4', 'HDCM_uv5_gpa', 
            'CBHD_1-họ tên', 'CBHD_1_C1.1', 'CBHD_1_C1.2', 'CBHD_1_C5.1', 'CBHD_1_gpa', 
            'CBHD_2-họ tên', 'CBHD_2_C2.1', 'CBHD_2_C2.2', 'CBHD_2_C3.1', 'CBHD_2_C5.2', 'CBHD_2_gpa', 
            'CBHD_3-họ tên', 'CBHD_3_C2.3', 'CBHD_3_C3.2', 'CBHD_3_C4.1', 'CBHD_3_C6.1', 'CBHD_3_C6.2', 'CBHD_3_gpa', 
            'CBPB-họ tên', 'CBPB_C2.3', 'CBPB_C3.2', 'CBPB_C4.1', 'CBPB_C6.1', 'CBPB_C6.2', 'CBPB_gpa'
            ]
        

        # Duyệt qua từng sinh viên
        for student in students:
            msv = student['msv']
            student_found = False
            for row in sheet.iter_rows(min_row=2, max_col=len(headers), values_only=False):
                cell_msv = row[headers['mã sinh viên'] - 1]
                # Nếu đã có msv, kiểm tra cột CBHD_1-họ tên trống để ghi
                if cell_msv.value == msv:
                    student_found = True
                    cell_cbhd_name = row[headers['cbpb-họ tên'] - 1]
                    if not cell_cbhd_name.value:  
                        cell_cbhd_name.value = lecturer_name
                        row[headers['cbpb_c2.3'] - 1].value = student['diemC23']
                        row[headers['cbpb_c3.2'] - 1].value = student['diemC32']
                        row[headers['cbpb_c4.1'] - 1].value = student['diemC41']
                        row[headers['cbpb_c6.1'] - 1].value = student['diemC61']
                        row[headers['cbpb_c6.2'] - 1].value = student['diemC62']
                        row[headers['cbpb_gpa'] - 1].value = student['gpa']
                    break

            # Nếu chưa có msv trong file
            if not student_found:
                # Tạo dòng mới
                new_row = [
                    student['fullname'],        
                    student['msv'],            
                    student['class'],             
                ] +[""]*53 + [
                    lecturer_name ,
                    student['diemC23'],         
                    student['diemC32'],         
                    student['diemC41'],
                    student['diemC61'],
                    student['diemC62'],
                    student['gpa']
                ]
                sheet.append(new_row)

        # Lưu file Excel
        workbook.save(file_path)
        xuat()
        # Chuyển hướng đến trang testOutput.html
        return JsonResponse({'message': 'Chấm điểm thành công'})

    # Nếu không phải POST, chuyển về baoCaoTienDoL1
    return JsonResponse({'message': 'Chấm điểm không thành công'})

def getGradeOfStudent(msv, name, formType):
    grade_student = {
    }
    file_path = os.path.join('DataCollected', 'final_new.xlsx')
    df = pd.read_excel(file_path)
    df_sv = df[df['Mã sinh viên'] == msv]
    if df_sv.empty:
        return None
    prefix_column = 'ZZZ'
    formTypeToPrefix = {
        'baoCaoTienDoL1': 'CBHD_1',
        'baoCaoTienDoL2': 'CBHD_2',
        'huongdan3': 'CBHD_3',
        'canBoPhanBien': 'CBPB',
        'hoiDongChuyenMon': 'HDCM'
    }
    for column in df_sv.columns:
        if df_sv[column].values[0] == name:
            column_tmp = column.split('-')
            if formTypeToPrefix[formType] in column_tmp[0]:
                prefix_column = column_tmp[0]
                break
    grade_columns = [
            'HDCM_uv1_C3.3', 'HDCM_uv1_C4.2', 'HDCM_uv1_C5.3', 'HDCM_uv1_C6.3', 'HDCM_uv1_C6.4', 'HDCM_uv1_gpa', 
            'HDCM_uv2_C3.3', 'HDCM_uv2_C4.2', 'HDCM_uv2_C5.3', 'HDCM_uv2_C6.3', 'HDCM_uv2_C6.4', 'HDCM_uv2_gpa', 
            'HDCM_uv3_C3.3', 'HDCM_uv3_C4.2', 'HDCM_uv3_C5.3', 'HDCM_uv3_C6.3', 'HDCM_uv3_C6.4', 'HDCM_uv3_gpa', 
            'HDCM_uv4_C3.3', 'HDCM_uv4_C4.2', 'HDCM_uv4_C5.3', 'HDCM_uv4_C6.3', 'HDCM_uv4_C6.4', 'HDCM_uv4_gpa', 
            'HDCM_uv5_C3.3', 'HDCM_uv5_C4.2', 'HDCM_uv5_C5.3', 'HDCM_uv5_C6.3', 'HDCM_uv5_C6.4', 'HDCM_uv5_gpa', 
            'CBHD_1_C1.1', 'CBHD_1_C1.2', 'CBHD_1_C5.1', 'CBHD_1_gpa', 
            'CBHD_2_C2.1', 'CBHD_2_C2.2', 'CBHD_2_C3.1', 'CBHD_2_C5.2', 'CBHD_2_gpa', 
            'CBHD_3_C2.3', 'CBHD_3_C3.2', 'CBHD_3_C4.1', 'CBHD_3_C6.1', 'CBHD_3_C6.2', 'CBHD_3_gpa', 
           'CBPB_C2.3', 'CBPB_C3.2', 'CBPB_C4.1', 'CBPB_C6.1', 'CBPB_C6.2', 'CBPB_gpa'
    ]
    true_columns = []
    for column in grade_columns:
        if column.startswith(prefix_column):
            true_columns.append(column)
    for column in true_columns:
        if not pd.isnull(df_sv[column].values[0]):
            grade_student[column.split('_')[-1].replace(".", "_")] = df_sv[column].values[0]
    return grade_student

def process_final_new_baocao1():
    # Đọc file Excel
    final_new_path = r"DataCollected\final_new.xlsx"
    tong_hop_diem_path = r"DataCollected\TongHopDiem1.xlsx"

    df_final_new = pd.read_excel(final_new_path)
    df_tong_hop_diem = pd.read_excel(tong_hop_diem_path)
    
    # Điền giá trị NaN bằng 0
    df_final_new = df_final_new.fillna(0)
    
    # Danh sách các cột cần lấy dữ liệu
    columns_to_extract = [
        'Họ và tên', 'Mã sinh viên', 'Lớp',
        'HDCM_uv1_C3', 'HDCM_uv1_C4', 'HDCM_uv1_C5', 'HDCM_uv1_C6',
        'HDCM_uv2_C3', 'HDCM_uv2_C4', 'HDCM_uv2_C5', 'HDCM_uv2_C6',
        'HDCM_uv3_C3', 'HDCM_uv3_C4', 'HDCM_uv3_C5', 'HDCM_uv3_C6',
        'HDCM_uv4_C3', 'HDCM_uv4_C4', 'HDCM_uv4_C5', 'HDCM_uv4_C6',
        'HDCM_uv5_C3', 'HDCM_uv5_C4', 'HDCM_uv5_C5', 'HDCM_uv5_C6',
        'CBHD_1_C1', 'CBHD_1_C5',
        'CBHD_2_C2', 'CBHD_2_C3', 'CBHD_2_C5',
        'CBHD_3_C2', 'CBHD_3_C3', 'CBHD_3_C4', 'CBHD_3_C6',
        'CBPB_C2', 'CBPB_C3', 'CBPB_C4', 'CBPB_C6',
        'o1TB', 'o2TB', 'o3TB', 'o4TB', 'o5TB', 'o6TB', 'gpa'
    ]
    
    # Hàm tính trung bình của các cột con
    def calculate_average(df, cols):
        existing_cols = [col for col in cols if col in df.columns]
        if existing_cols:
            return df[existing_cols].mean(axis=1)
        else:
            return pd.Series([0] * len(df))
    
    # Tạo DataFrame mới với các cột cần thiết
    new_data = pd.DataFrame()
    new_data['Họ và tên'] = df_final_new['Họ và tên']
    new_data['Mã sinh viên'] = df_final_new['Mã sinh viên']
    new_data['Lớp'] = df_final_new['Lớp']
    
    # Tính trung bình cho các cột HDCM_uv
    for i in range(1, 6):
        for j in range(3, 7):
            col_name = f'HDCM_uv{i}_C{j}'
            sub_cols = [f'HDCM_uv{i}_C{j}.{k}' for k in range(1, 5) if f'HDCM_uv{i}_C{j}.{k}' in df_final_new.columns]
            new_data[col_name] = calculate_average(df_final_new, sub_cols)
    
    # Tính trung bình cho các cột CBHD
    for i in range(1, 4):
        for j in [1, 2, 3, 4, 5, 6]:
            col_name = f'CBHD_{i}_C{j}'
            sub_cols = [f'CBHD_{i}_C{j}.{k}' for k in range(1, 5) if f'CBHD_{i}_C{j}.{k}' in df_final_new.columns] 
            new_data[col_name] = calculate_average(df_final_new, sub_cols)
    
    # Tính trung bình cho các cột CBPB
    for j in [2, 3, 4, 6]:
        col_name = f'CBPB_C{j}'
        sub_cols = [f'CBPB_C{j}.{k}' for k in range(1, 5) if f'CBPB_C{j}.{k}' in df_final_new.columns]
        new_data[col_name] = calculate_average(df_final_new, sub_cols)
    
    # Tính trung bình cho các cột o1TB, o2TB, ..., o6TB
    for i in range(1, 7):
        col_name = f'o{i}TB'
        sub_cols = [f'HDCM_uv{j}_C{i}' for j in range(1, 6) if f'HDCM_uv{j}_C{i}' in columns_to_extract] + \
           [f'CBHD_{j}_C{i}' for j in range(1, 4) if f'CBHD_{j}_C{i}' in columns_to_extract] + \
           [f'CBPB_C{i}' for j in range(1, 2) if f'CBPB_C{i}' in columns_to_extract]
        new_data[col_name] = calculate_average(new_data, sub_cols)
        # print(sub_cols)
        # print(df_final_new.columns)

    # Tính trung bình cho cột gpa
    new_data['gpa'] = new_data[[f'o{i}TB' for i in range(1, 7)]].mean(axis=1)
    
    # Chỉ giữ lại các cột cần thiết
    new_data = new_data[columns_to_extract]

    # Xóa dữ liệu cũ từ hàng số 4 trở đi
    workbook = load_workbook(tong_hop_diem_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

    # Lưu workbook sau khi xóa dữ liệu cũ
    workbook.save(tong_hop_diem_path)
    
    # Thêm dữ liệu mới vào file TongHopDiem1.xlsx
    with pd.ExcelWriter(tong_hop_diem_path, mode='a', if_sheet_exists='overlay') as writer:
        new_data.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
    
    print("Dữ liệu đã được thêm vào file TongHopDiem1.xlsx")

def process_final_new_baocao2():
    # Đọc file Excel
    final_new_path = r"DataCollected\final_new.xlsx"
    df_final_new = pd.read_excel(final_new_path)
    
    # Điền giá trị NaN bằng 0
    df_final_new = df_final_new.fillna(0)
    
    # Danh sách các cột cần lấy dữ liệu
    columns_to_extract = [
        'TT',
        'Họ và tên', 'Mã sinh viên', 'Lớp',
        'CBHD_1_C1', 'CBHD_1_C5', 'CBHD_1_gpa',
        'CBHD_2_C2', 'CBHD_2_C3', 'CBHD_2_C5', 'CBHD_2_gpa',
        'o1TB', 'o2TB', 'o3TB', 'o4TB', 'o5TB', 'o6TB', 'oGPA',
        'o2CK', 'o3CK', 'o4CK', 'o5CK', 'o6CK', 'GPA_CK',
        'GPA_tong'
    ]
    
    # Hàm tính trung bình của các cột con
    def calculate_average(df, cols):
        existing_cols = [col for col in cols if col in df.columns]
        if existing_cols:
            return df[existing_cols].mean(axis=1)
        else:
            return pd.Series([0] * len(df))
    
    # Tạo DataFrame mới với các cột cần thiết
    new_data = pd.DataFrame()
    new_data['TT'] = range(1, len(df_final_new) + 1)
    new_data['Họ và tên'] = df_final_new['Họ và tên']
    new_data['Mã sinh viên'] = df_final_new['Mã sinh viên']
    new_data['Lớp'] = df_final_new['Lớp']
    
    # Tính trung bình cho các cột CBHD_1_C1, CBHD_1_C5, CBHD_2_C2, CBHD_2_C3, CBHD_2_C5
    for i in range(1, 3):
        for j in [1, 2, 3, 5]:
            if i == 1 and j == 2:
                continue
            col_name = f'CBHD_{i}_C{j}'
            sub_cols = [f'CBHD_{i}_C{j}.{k}' for k in range(1, 5) if f'CBHD_{i}_C{j}.{k}' in df_final_new.columns]
            new_data[col_name] = calculate_average(df_final_new, sub_cols)
    
    # Thêm các cột CBHD_1_gpa và CBHD_2_gpa
    new_data['CBHD_1_gpa'] = df_final_new['CBHD_1_gpa']
    new_data['CBHD_2_gpa'] = df_final_new['CBHD_2_gpa']
    
    # Tính trung bình cho các cột o1TB, o2TB, ..., o6TB
    for i in range(1, 7):
        col_name = f'o{i}TB'
        sub_cols = [f'HDCM_uv{j}_C{i}.{k}' for j in range(1, 6) for k in range(1, 5) if f'HDCM_uv{j}_C{i}.{k}' in df_final_new.columns] + \
                   [f'CBHD_{j}_C{i}.{k}' for j in range(1, 4) for k in range(1, 5) if f'CBHD_{j}_C{i}.{k}' in df_final_new.columns] + \
                   [f'CBPB_C{i}.{k}' for k in range(1, 5) if f'CBPB_C{i}.{k}' in df_final_new.columns]
        new_data[col_name] = calculate_average(df_final_new, sub_cols)
    
    # Tính trung bình cho cột oGPA
    new_data['oGPA'] = new_data[[f'o{i}TB' for i in range(1, 7)]].mean(axis=1)
    
    # Tính trung bình cho các cột o2CK, o3CK, ..., o6CK
    for i in range(2, 7):
        col_name = f'o{i}CK'
        sub_cols = [f'HDCM_uv{j}_C{i}.{k}' for j in range(1, 6) for k in range(1, 5) if f'HDCM_uv{j}_C{i}.{k}' in df_final_new.columns] + \
                   [f'CBHD_{j}_C{i}.{k}' for j in range(3,4) for k in range(1, 5) if f'CBHD_{j}_C{i}.{k}' in df_final_new.columns] + \
                   [f'CBPB_C{i}.{k}' for k in range(1, 5) if f'CBPB_C{i}.{k}' in df_final_new.columns]
        new_data[col_name] = calculate_average(df_final_new, sub_cols)
    
    # Tính trung bình cho cột GPA_CK
    new_data['GPA_CK'] = new_data[[f'o{i}CK' for i in range(2, 7)]].mean(axis=1)
    
    # Tính trung bình cho cột GPA_tong
    new_data['GPA_tong'] = new_data['CBHD_1_gpa'] * 0.1 + new_data['CBHD_2_gpa'] * 0.2 + new_data['oGPA'] * 0.35 + new_data['GPA_CK'] * 0.35
    
    # Chỉ giữ lại các cột cần thiết
    new_data = new_data[columns_to_extract]

    # Đường dẫn đến file TongHopDiem2.xlsx
    tong_hop_diem2_path = r"DataCollected\TongHopDiem2.xlsx"

    # Xóa dữ liệu cũ từ hàng số 4 trở đi và thêm dữ liệu mới
    workbook = load_workbook(tong_hop_diem2_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

    # Lưu workbook sau khi xóa dữ liệu cũ
    workbook.save(tong_hop_diem2_path)

    # Thêm dữ liệu mới vào file TongHopDiem2.xlsx
    with pd.ExcelWriter(tong_hop_diem2_path, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
        new_data.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=3)

    print("Dữ liệu đã được thêm vào file TongHopDiem2.xlsx")
 
def xuat():
    process_final_new_baocao1()
    process_final_new_baocao2()
